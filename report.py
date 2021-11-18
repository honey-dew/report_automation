import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment, numbers
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.formatting.rule import IconSet, FormatObject, Rule

df = pd.read_excel("data/data_input.xlsx")

# Add column Order Year
df["Order Year"] = df["Order Date"].dt.year.astype("str")

# Pivoting by Country as index, Order Year as column, and Sales as value
pv_country = pd.pivot_table(
    data = df,
    index = "Country",
    columns = "Order Year",
    values = "Sales",
    aggfunc = "sum"
).reset_index()

def add_column_growth(input_df):
    """
    Function to calculate growth from 2011 to 2014
    """
    df = input_df
    input_df["Growth"] = ((input_df["2014"] / input_df["2011"])**(1/4))-1
    return df

# add_column_growth(pv_sub_category)
pv_country["Growth"] = (((pv_country["2014"] / pv_country["2011"])**(1/4))-1)
pv_country = pv_country.sort_values(by="Growth", ascending=False)

# Creating Workbook
wb = Workbook()
wb.active.title = "report"

# Creating tuple from pv_country
data_pv_country = pv_country.to_records(index=False).tolist()

ws = wb["report"]
ws["A2"].value = "Sales Growth per Country"
ws["A2"].font = Font(size=12, bold=True)
ws.append(pv_country.columns.tolist())
for row in data_pv_country:
    ws.append(row)

ws[f'{get_column_letter(ws.min_column)}{ws.max_row + 1}'].value = "Total"
ws[f'{get_column_letter(ws.min_column)}{ws.max_row}'].font = Font(bold=True)

# Get Total
for col in range(ws.min_column+1, ws.max_column):
    ws[f'{get_column_letter(col)}{ws.max_row}'].value = f'=SUM({get_column_letter(col)}{ws.min_row+2}:{get_column_letter(col)}{ws.max_row-1})'

# Get total growth
# ((present / past)**(1/4))-1
min_row_past = f'{get_column_letter(ws.min_column+1)}{ws.min_row+2}'
max_row_past = f'{get_column_letter(ws.min_column+1)}{ws.max_row-1}'
min_row_present = f'{get_column_letter(ws.max_column-1)}{ws.min_row+2}'
max_row_present = f'{get_column_letter(ws.max_column-1)}{ws.max_row-1}'
past = f'SUM({min_row_past}:{max_row_past})'
present = f'SUM({min_row_present}:{max_row_present})'
period = (ws.max_column - ws.min_column) - 1
total_growth = f'=(({present}/{past})^({1/period})-1)'
# ws[f'{get_column_letter(ws.max_column)}{ws.max_row}'].value = "=((E19/B19)^(1/4))-1"
ws[f'{get_column_letter(ws.max_column)}{ws.max_row}'].value = total_growth

# Adding border and change font size
for row in range(ws.min_row+1, ws.max_row+1):
    for col in range(ws.max_column):
        border_style = Side(border_style="thin", color="fcf3cf")
        cell = ws.cell(column=col+1, row=row)
        cell.border = Border(right=border_style, top=border_style, bottom=border_style, left=border_style)
        cell.font = Font(size=10)

# Styling header
for col in range(ws.max_column):
    cell = ws.cell(column=col+1, row=ws.min_row+1)
    cell.font = Font(bold=True, color="fcf3cf")
    cell.fill = PatternFill(fgColor="229954", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Styling Total row
for col in range(ws.min_column, ws.max_column+1):
    ws[f'{get_column_letter(col)}{ws.max_row}'].font = Font(bold=True, size=10)
    border_style = Side(border_style="thick", color="229954")
    ws[f'{get_column_letter(col)}{ws.max_row}'].border = Border(top=border_style)

def get_style_number(sheet_obj, column_letter, format_number):
    """
    Function to style format number each row
    """
    for row in range(sheet_obj.max_row):
        cell = ws.cell(column=column_index_from_string(column_letter), row=row+1)
        cell.number_format = format_number

# Styling Growth column
get_style_number(ws, "F", numbers.FORMAT_PERCENTAGE)

# Styling year column (column B to E)
for col in range(2,6):
    get_style_number(ws, get_column_letter(col), numbers.FORMAT_NUMBER_COMMA_SEPARATED1)

# Adding conditional formatting
first = FormatObject(type='num', val=-100000)
second = FormatObject(type='num', val=0)
third = FormatObject(type='num', val=0.005)
iconset = IconSet(iconSet='3Arrows', cfvo=[first, second, third], showValue=None, percent=None, reverse=None)
rule = Rule(type='iconSet', iconSet=iconset)
ws.conditional_formatting.add("F4:F19", rule)

for col in range(1, ws.max_column+1):
    ws.column_dimensions[get_column_letter(col)].autosize = True

wb.save("data/data_output.xlsx")
wb.close()